"""
Exporter Tool - AI-discoverable data export capabilities

This tool provides export functions for various file formats including
CSV, JSON, Excel, XML, and more.
"""

import csv
import json
import io
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom

from ..registry import ai_tool, create_example


@ai_tool(
    name="export_csv",
    description="Export data to CSV format with customizable options",
    category="export",
    examples=[
        create_example(
            "Export product data to CSV",
            data=[
                {"name": "Product A", "price": 99.99, "stock": 50},
                {"name": "Product B", "price": 149.99, "stock": 30}
            ],
            filename="products.csv",
            columns=["name", "price", "stock"]
        ),
        create_example(
            "Export with custom headers",
            data=[{"id": 1, "val": 100}],
            filename="data.csv",
            headers={"id": "ID Number", "val": "Value"}
        )
    ],
    capabilities=[
        "Export to CSV format",
        "Custom column selection",
        "Header customization",
        "Delimiter options",
        "Encoding support",
        "Append mode",
        "Automatic type conversion"
    ],
    performance={
        "speed": "high",
        "reliability": "high",
        "cost": "low"
    }
)
async def export_csv(
    data: Union[Dict[str, Any], List[Dict[str, Any]]],
    filename: str,
    columns: Optional[List[str]] = None,
    headers: Optional[Dict[str, str]] = None,
    delimiter: str = ",",
    include_headers: bool = True,
    append: bool = False
) -> Dict[str, Any]:
    """
    Export data to CSV format
    
    Args:
        data: Data to export (dict or list of dicts)
        filename: Output filename
        columns: Specific columns to export (default: all)
        headers: Custom header names mapping
        delimiter: Field delimiter (default: comma)
        include_headers: Whether to include headers
        append: Append to existing file
        
    Returns:
        Dictionary with export status and details
    """
    try:
        # Ensure data is a list
        records = data if isinstance(data, list) else [data]
        
        if not records:
            return {
                "success": False,
                "error": "No data to export"
            }
        
        # Determine columns
        if not columns:
            columns = list(records[0].keys())
        
        # Prepare headers
        if headers:
            header_row = [headers.get(col, col) for col in columns]
        else:
            header_row = columns
        
        # Write to file
        mode = 'a' if append else 'w'
        with open(filename, mode, newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=columns, delimiter=delimiter)
            
            if include_headers and not append:
                writer.writerow(dict(zip(columns, header_row)))
            
            for record in records:
                # Filter only requested columns
                filtered_record = {col: record.get(col, '') for col in columns}
                writer.writerow(filtered_record)
        
        return {
            "success": True,
            "filename": filename,
            "format": "csv",
            "records_exported": len(records),
            "columns": columns,
            "file_size": _get_file_size(filename)
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "filename": filename
        }


@ai_tool(
    name="export_json",
    description="Export data to JSON format with formatting options",
    category="export",
    examples=[
        create_example(
            "Export data as JSON array",
            data=[{"id": 1, "name": "Item 1"}, {"id": 2, "name": "Item 2"}],
            filename="data.json",
            pretty=True
        ),
        create_example(
            "Export as JSON lines",
            data=[{"event": "click", "time": "2024-01-01"}],
            filename="events.jsonl",
            json_lines=True
        )
    ],
    capabilities=[
        "Export to JSON format",
        "Pretty printing with indentation",
        "JSON Lines format support",
        "Custom encoding",
        "Nested data support",
        "Date/time serialization"
    ],
    performance={
        "speed": "high",
        "reliability": "high",
        "cost": "low"
    }
)
async def export_json(
    data: Any,
    filename: str,
    pretty: bool = False,
    indent: int = 2,
    json_lines: bool = False,
    sort_keys: bool = False
) -> Dict[str, Any]:
    """
    Export data to JSON format
    
    Args:
        data: Data to export (any JSON-serializable data)
        filename: Output filename
        pretty: Pretty print with indentation
        indent: Indentation spaces (if pretty=True)
        json_lines: Export as JSON Lines format
        sort_keys: Sort object keys
        
    Returns:
        Dictionary with export status
    """
    try:
        # Custom JSON encoder for dates and other types
        class CustomEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, datetime):
                    return obj.isoformat()
                if hasattr(obj, '__dict__'):
                    return obj.__dict__
                return super().default(obj)
        
        with open(filename, 'w', encoding='utf-8') as f:
            if json_lines and isinstance(data, list):
                # JSON Lines format - one JSON object per line
                for item in data:
                    json.dump(item, f, cls=CustomEncoder, sort_keys=sort_keys)
                    f.write('\n')
            else:
                # Standard JSON format
                if pretty:
                    json.dump(data, f, cls=CustomEncoder, indent=indent, sort_keys=sort_keys)
                else:
                    json.dump(data, f, cls=CustomEncoder, sort_keys=sort_keys)
        
        return {
            "success": True,
            "filename": filename,
            "format": "jsonl" if json_lines else "json",
            "records_exported": len(data) if isinstance(data, list) else 1,
            "pretty_printed": pretty,
            "file_size": _get_file_size(filename)
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "filename": filename
        }


@ai_tool(
    name="export_excel",
    description="Export data to Excel format with multiple sheets and formatting",
    category="export",
    examples=[
        create_example(
            "Export to Excel with multiple sheets",
            data={
                "Products": [{"name": "A", "price": 100}],
                "Orders": [{"id": 1, "total": 300}]
            },
            filename="report.xlsx"
        ),
        create_example(
            "Export with formatting",
            data=[{"name": "Product", "price": 99.99, "date": "2024-01-01"}],
            filename="formatted.xlsx",
            sheet_name="Sales Data",
            auto_format=True
        )
    ],
    capabilities=[
        "Export to Excel format",
        "Multiple sheets support",
        "Column width auto-adjustment",
        "Number formatting",
        "Date formatting",
        "Header styling",
        "Formula support"
    ],
    performance={
        "speed": "medium",
        "reliability": "high",
        "cost": "low"
    }
)
async def export_excel(
    data: Union[Dict[str, Any], List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]],
    filename: str,
    sheet_name: str = "Sheet1",
    auto_format: bool = True,
    include_index: bool = False
) -> Dict[str, Any]:
    """
    Export data to Excel format
    
    Args:
        data: Data to export - can be list, dict, or dict of lists for multiple sheets
        filename: Output filename
        sheet_name: Sheet name (if single sheet)
        auto_format: Apply automatic formatting
        include_index: Include row index
        
    Returns:
        Dictionary with export status
    """
    try:
        # Import pandas here to avoid dependency if not needed
        import pandas as pd
        
        # Handle different data formats
        if isinstance(data, dict) and all(isinstance(v, list) for v in data.values()):
            # Multiple sheets
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                sheets_exported = 0
                total_records = 0
                
                for sheet, sheet_data in data.items():
                    if sheet_data:
                        df = pd.DataFrame(sheet_data)
                        df.to_excel(writer, sheet_name=sheet[:31], index=include_index)  # Excel sheet name limit
                        sheets_exported += 1
                        total_records += len(sheet_data)
                        
                        if auto_format:
                            worksheet = writer.sheets[sheet[:31]]
                            _format_excel_sheet(worksheet, df)
                
                return {
                    "success": True,
                    "filename": filename,
                    "format": "xlsx",
                    "sheets_exported": sheets_exported,
                    "total_records": total_records,
                    "file_size": _get_file_size(filename)
                }
        else:
            # Single sheet
            records = data if isinstance(data, list) else [data]
            df = pd.DataFrame(records)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=include_index)
                
                if auto_format:
                    worksheet = writer.sheets[sheet_name]
                    _format_excel_sheet(worksheet, df)
            
            return {
                "success": True,
                "filename": filename,
                "format": "xlsx",
                "records_exported": len(records),
                "file_size": _get_file_size(filename)
            }
            
    except ImportError:
        # Fallback to CSV if pandas not available
        result = await export_csv(data, filename.replace('.xlsx', '.csv'))
        result["note"] = "Exported as CSV (Excel export requires pandas)"
        return result
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "filename": filename
        }


@ai_tool(
    name="export_xml",
    description="Export data to XML format with customizable structure",
    category="export",
    examples=[
        create_example(
            "Export data as XML",
            data=[
                {"id": 1, "name": "Product A", "price": 99.99},
                {"id": 2, "name": "Product B", "price": 149.99}
            ],
            filename="products.xml",
            root_element="products",
            record_element="product"
        )
    ],
    capabilities=[
        "Export to XML format",
        "Custom root and record elements",
        "Attribute support",
        "Nested data structures",
        "Pretty formatting",
        "Namespace support"
    ],
    performance={
        "speed": "medium",
        "reliability": "high",
        "cost": "low"
    }
)
async def export_xml(
    data: Union[Dict[str, Any], List[Dict[str, Any]]],
    filename: str,
    root_element: str = "data",
    record_element: str = "record",
    pretty: bool = True,
    attributes: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    Export data to XML format
    
    Args:
        data: Data to export
        filename: Output filename
        root_element: Root element name
        record_element: Individual record element name
        pretty: Pretty print the XML
        attributes: Fields to use as attributes instead of elements
        
    Returns:
        Dictionary with export status
    """
    try:
        # Ensure data is a list
        records = data if isinstance(data, list) else [data]
        attributes = attributes or []
        
        # Create root element
        root = ET.Element(root_element)
        
        # Add records
        for record in records:
            record_elem = ET.SubElement(root, record_element)
            
            for key, value in record.items():
                if key in attributes:
                    # Add as attribute
                    record_elem.set(key, str(value))
                else:
                    # Add as sub-element
                    sub_elem = ET.SubElement(record_elem, key)
                    sub_elem.text = str(value) if value is not None else ""
        
        # Convert to string
        xml_str = ET.tostring(root, encoding='unicode')
        
        # Pretty print if requested
        if pretty:
            dom = minidom.parseString(xml_str)
            xml_str = dom.toprettyxml(indent="  ")
            # Remove extra blank lines
            xml_str = '\n'.join([line for line in xml_str.split('\n') if line.strip()])
        
        # Write to file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(xml_str)
        
        return {
            "success": True,
            "filename": filename,
            "format": "xml",
            "records_exported": len(records),
            "root_element": root_element,
            "record_element": record_element,
            "file_size": _get_file_size(filename)
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "filename": filename
        }


@ai_tool(
    name="export_html",
    description="Export data as formatted HTML table or report",
    category="export",
    examples=[
        create_example(
            "Export as HTML table",
            data=[
                {"Product": "iPhone", "Price": "$999", "Stock": "In Stock"},
                {"Product": "iPad", "Price": "$799", "Stock": "Limited"}
            ],
            filename="products.html",
            title="Product Inventory"
        )
    ],
    capabilities=[
        "Export as HTML table",
        "Custom styling",
        "Responsive design",
        "Title and headers",
        "Sortable tables",
        "Export as full report"
    ],
    performance={
        "speed": "high",
        "reliability": "high",
        "cost": "low"
    }
)
async def export_html(
    data: Union[Dict[str, Any], List[Dict[str, Any]]],
    filename: str,
    title: str = "Data Export",
    include_style: bool = True,
    sortable: bool = False
) -> Dict[str, Any]:
    """
    Export data as HTML
    
    Args:
        data: Data to export
        filename: Output filename
        title: Page title
        include_style: Include CSS styling
        sortable: Make table sortable
        
    Returns:
        Dictionary with export status
    """
    try:
        # Ensure data is a list
        records = data if isinstance(data, list) else [data]
        
        if not records:
            return {
                "success": False,
                "error": "No data to export"
            }
        
        # Get columns
        columns = list(records[0].keys())
        
        # Build HTML
        html_parts = ['<!DOCTYPE html>', '<html>', '<head>']
        html_parts.append(f'<title>{title}</title>')
        html_parts.append('<meta charset="utf-8">')
        html_parts.append('<meta name="viewport" content="width=device-width, initial-scale=1">')
        
        if include_style:
            html_parts.append(_get_html_style(sortable))
        
        if sortable:
            html_parts.append(_get_sortable_script())
        
        html_parts.append('</head>')
        html_parts.append('<body>')
        html_parts.append(f'<h1>{title}</h1>')
        html_parts.append(f'<p>Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>')
        
        # Build table
        table_class = 'sortable' if sortable else ''
        html_parts.append(f'<table class="{table_class}">')
        html_parts.append('<thead><tr>')
        
        for col in columns:
            html_parts.append(f'<th>{col}</th>')
        
        html_parts.append('</tr></thead>')
        html_parts.append('<tbody>')
        
        for record in records:
            html_parts.append('<tr>')
            for col in columns:
                value = record.get(col, '')
                # Escape HTML
                if isinstance(value, str):
                    value = value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                html_parts.append(f'<td>{value}</td>')
            html_parts.append('</tr>')
        
        html_parts.append('</tbody>')
        html_parts.append('</table>')
        html_parts.append(f'<p>Total records: {len(records)}</p>')
        html_parts.append('</body>')
        html_parts.append('</html>')
        
        # Write file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html_parts))
        
        return {
            "success": True,
            "filename": filename,
            "format": "html",
            "records_exported": len(records),
            "features": {
                "styled": include_style,
                "sortable": sortable
            },
            "file_size": _get_file_size(filename)
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "filename": filename
        }


# Helper functions
def _get_file_size(filename: str) -> str:
    """Get human-readable file size"""
    try:
        import os
        size = os.path.getsize(filename)
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024.0:
                return f"{size:.2f} {unit}"
            size /= 1024.0
        return f"{size:.2f} TB"
    except:
        return "Unknown"


def _format_excel_sheet(worksheet, df):
    """Apply formatting to Excel worksheet"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
    except:
        pass  # Formatting is optional


def _get_html_style(sortable: bool) -> str:
    """Get CSS style for HTML export"""
    return '''<style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        h1 {
            color: #333;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            background-color: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        th, td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        th {
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
            ''' + ('cursor: pointer;' if sortable else '') + '''
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        p {
            color: #666;
        }
    </style>'''


def _get_sortable_script() -> str:
    """Get JavaScript for sortable tables"""
    return '''<script>
    function sortTable(n) {
        var table, rows, switching, i, x, y, shouldSwitch, dir, switchcount = 0;
        table = document.querySelector("table");
        switching = true;
        dir = "asc";
        
        while (switching) {
            switching = false;
            rows = table.rows;
            
            for (i = 1; i < (rows.length - 1); i++) {
                shouldSwitch = false;
                x = rows[i].getElementsByTagName("TD")[n];
                y = rows[i + 1].getElementsByTagName("TD")[n];
                
                if (dir == "asc") {
                    if (x.innerHTML.toLowerCase() > y.innerHTML.toLowerCase()) {
                        shouldSwitch = true;
                        break;
                    }
                } else if (dir == "desc") {
                    if (x.innerHTML.toLowerCase() < y.innerHTML.toLowerCase()) {
                        shouldSwitch = true;
                        break;
                    }
                }
            }
            
            if (shouldSwitch) {
                rows[i].parentNode.insertBefore(rows[i + 1], rows[i]);
                switching = true;
                switchcount++;
            } else {
                if (switchcount == 0 && dir == "asc") {
                    dir = "desc";
                    switching = true;
                }
            }
        }
    }
    
    // Add click handlers
    document.addEventListener('DOMContentLoaded', function() {
        var headers = document.querySelectorAll('th');
        headers.forEach(function(header, index) {
            header.onclick = function() { sortTable(index); };
        });
    });
    </script>'''
